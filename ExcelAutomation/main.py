#source .venv/bin/activate
#changes the name of the directory
import openpyxl as xl
#from the package you have a module called chart and from the module you're importing two classes
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook ('transactions2.xlsx')
    #wb is a directory and sheet is being set = to fileName
sheet = wb ['Sheet1']
    #cell = sheet.cell(1, 1)


    #range only goes from the first value to the second value not including the last of the second value
for row in range (2, sheet.max_row + 1):
    cell = sheet.cell (row, 3)
    print(cell.value)
    corrected_price = cell.value * 0.5
     #creates a new cell in the 4th column, or accesses that particular cell
    corrected_price_cell = sheet.cell (row, 4)
    #sets the value of the new cell to the corrected price
    corrected_price_cell.value = corrected_price

values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col = 4, max_col=4)
chart = BarChart ()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save ('transactions2.xlsx')
